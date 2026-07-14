"""
report_generator.py — Excel Report Generator
Oral Ulcer AI — Saveetha Dental College
Surgical Luxury Theme (Maroon #7B1E3A / Gold #C9A84C)
"""
import os
import sys
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def generate_report(results, filename="Web_Test_Execution_Report.xlsx"):
    """
    Generates a beautifully styled Excel spreadsheet containing test run results.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Web Test Execution"

    # ── Color Palette ─────────────────────────────────────────────
    MAROON      = "7B1E3A"
    DARK_MAROON = "5C1028"
    GOLD        = "C9A84C"
    IVORY       = "FAF7F4"
    PASS_BG     = "E8F5E9"
    FAIL_BG     = "FFEBEE"
    PASS_FG     = "1B5E20"
    FAIL_FG     = "B71C1C"
    STRIPE_BG   = "F5F0ED"

    # ── Styles ────────────────────────────────────────────────────
    maroon_fill  = PatternFill("solid", fgColor=MAROON)
    stripe_fill  = PatternFill("solid", fgColor=STRIPE_BG)
    ivory_fill   = PatternFill("solid", fgColor=IVORY)
    pass_fill    = PatternFill("solid", fgColor=PASS_BG)
    fail_fill    = PatternFill("solid", fgColor=FAIL_BG)

    header_font  = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    title_font   = Font(name="Calibri", size=14, bold=True, color=MAROON)
    subtitle_font= Font(name="Calibri", size=9,  italic=True, color="9E8A8F")
    regular_font = Font(name="Calibri", size=9)
    pass_font    = Font(name="Calibri", size=9, bold=True, color=PASS_FG)
    fail_font    = Font(name="Calibri", size=9, bold=True, color=FAIL_FG)
    id_font      = Font(name="Calibri", size=9, bold=True, color=MAROON)
    gold_font    = Font(name="Calibri", size=9, bold=True, color="A07828")
    mod_font     = Font(name="Calibri", size=9, bold=True, color="4A4A4A")

    thin = Side(style="thin",   color="D8C8C0")
    med  = Side(style="medium", color=GOLD)
    data_border  = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_border= Border(
        left=Side(style="thin", color=DARK_MAROON),
        right=Side(style="thin", color=DARK_MAROON),
        top=Side(style="thin", color=DARK_MAROON),
        bottom=Side(style="medium", color=GOLD)
    )

    center  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ── Title Banner (Rows 1–3) ────────────────────────────────────
    ws.merge_cells("A1:H1")
    ws["A1"] = "ORAL ULCER AI — WEB APPLICATION E2E TEST REPORT"
    ws["A1"].font      = title_font
    ws["A1"].alignment = center
    ws["A1"].fill      = PatternFill("solid", fgColor="FAF7F4")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    now = datetime.datetime.now().strftime("%d %B %Y  |  %H:%M:%S")
    ws["A2"] = f"Saveetha Dental College & Hospital  •  Generated: {now}"
    ws["A2"].font      = subtitle_font
    ws["A2"].alignment = center
    ws.row_dimensions[2].height = 18

    ws.merge_cells("A3:H3")
    ws["A3"] = ""
    ws.row_dimensions[3].height = 6

    # ── Summary Stats Row (Row 4) ──────────────────────────────────
    total   = len(results)
    passed  = sum(1 for r in results if r.get("status") == "PASS")
    failed  = total - passed
    rate    = f"{round(passed / total * 100)}%" if total else "0%"
    modules = len(set(r.get("module", "") for r in results))

    stats = [
        ("TOTAL TCs", str(total),  MAROON),
        ("PASSED",    str(passed), "2E7D32"),
        ("FAILED",    str(failed), "C62828"),
        ("PASS RATE", rate,        "A07828"),
        ("MODULES",   str(modules), "37474F"),
    ]
    col_positions = [1, 2, 3, 4, 5]
    for (label, value, color), col in zip(stats, col_positions):
        label_cell = ws.cell(row=4, column=col)
        label_cell.value     = label
        label_cell.font      = Font(name="Calibri", size=8, bold=True, color="888888")
        label_cell.alignment = center
        label_cell.fill      = PatternFill("solid", fgColor="F0EAE6")

        val_cell = ws.cell(row=5, column=col)
        val_cell.value     = value
        val_cell.font      = Font(name="Calibri", size=18, bold=True, color=color)
        val_cell.alignment = center
        val_cell.fill      = PatternFill("solid", fgColor="F0EAE6")

    ws.row_dimensions[4].height = 16
    ws.row_dimensions[5].height = 34

    ws.merge_cells("F4:H4")
    ws.merge_cells("F5:H5")
    ws["F4"].value     = "AUTOMATED E2E TESTING"
    ws["F4"].font      = Font(name="Calibri", size=8, bold=True, color="888888")
    ws["F4"].alignment = center
    ws["F5"].value     = "Selenium WebDriver / Desktop Chrome"
    ws["F5"].font      = Font(name="Calibri", size=9, color="4A4A4A", italic=True)
    ws["F5"].alignment = center

    ws.row_dimensions[6].height = 8

    # ── Column Headers (Row 7) ────────────────────────────────────
    headers = [
        "Test Case ID", "Platform", "Test Type", "Module",
        "Test Description", "Status", "Duration (ms)",
        "Error / Output Logs"
    ]
    ws.append([""] * 8)  # row 7 placeholder
    ws.append(headers)   # row 8

    header_row_idx = 8
    ws.row_dimensions[header_row_idx].height = 28
    for col, cell in enumerate(ws[header_row_idx], 1):
        cell.font      = header_font
        cell.fill      = maroon_fill
        cell.alignment = center
        cell.border    = header_border

    # ── Data Rows ─────────────────────────────────────────────────
    prev_module = None
    for i, test in enumerate(results):
        row_data = [
            test.get("id", ""),
            test.get("platform", "Desktop Web"),
            test.get("test_type", "Functional"),
            test.get("module", ""),
            test.get("description", ""),
            test.get("status", ""),
            test.get("duration", 0),
            test.get("logs", "No errors. Flow completed successfully."),
        ]
        ws.append(row_data)
        row_idx    = ws.max_row
        is_stripe  = (i % 2 == 1)
        row_fill   = stripe_fill if is_stripe else ivory_fill
        status     = test.get("status", "")
        cur_module = test.get("module", "")

        ws.row_dimensions[row_idx].height = 22

        for col_idx, cell in enumerate(ws[row_idx], 1):
            cell.border    = data_border
            cell.alignment = left

            if col_idx == 1:  # ID
                cell.font      = id_font
                cell.alignment = center
                cell.fill      = row_fill
            elif col_idx == 2:  # Platform
                cell.font      = Font(name="Calibri", size=8, color="6A5560")
                cell.alignment = center
                cell.fill      = row_fill
            elif col_idx == 3:  # Test Type
                cell.font      = mod_font
                cell.alignment = center
                cell.fill      = row_fill
            elif col_idx == 4:  # Module
                cell.font  = mod_font
                cell.fill  = row_fill
                if cur_module != prev_module:
                    cell.font = Font(name="Calibri", size=9, bold=True, color=MAROON)
                prev_module = cur_module
            elif col_idx == 5:  # Description
                cell.font = regular_font
                cell.fill = row_fill
            elif col_idx == 6:  # Status
                cell.alignment = center
                if status == "PASS":
                    cell.font = pass_font
                    cell.fill = pass_fill
                else:
                    cell.font = fail_font
                    cell.fill = fail_fill
            elif col_idx == 7:  # Duration
                cell.alignment = center
                cell.font      = Font(name="Calibri", size=9, color="4A4A4A")
                cell.fill      = row_fill
            elif col_idx == 8:  # Logs
                cell.font = Font(name="Calibri", size=8, color="6A5560", italic=(status == "PASS"))
                cell.fill = row_fill

    # ── Column Widths ─────────────────────────────────────────────
    widths = {"A": 14, "B": 18, "C": 15, "D": 22, "E": 52, "F": 11, "G": 14, "H": 55}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    # ── Freeze Panes ──────────────────────────────────────────────
    ws.freeze_panes = "A9"

    # ── Save ──────────────────────────────────────────────────────
    reports_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "reports")
    os.makedirs(reports_dir, exist_ok=True)
    output_path = os.path.join(reports_dir, filename)
    wb.save(output_path)

    print("\n" + "=" * 58)
    print("  [SUCCESS] Test report generated successfully!")
    print(f"  Path   : {output_path}")
    print(f"  Total  : {total}  |  Passed: {passed}  |  Failed: {failed}  |  Rate: {rate}")
    print("=" * 58 + "\n")


# ── Dry-Run (mock data for 105 tests) ─────────────────────────────
if __name__ == "__main__":
    if "--dry" in sys.argv:
        import random
        print("Generating comprehensive mock web test report (105 test cases)...")

        MOCK_TESTS = [
            ("TC-WEB-001","Splash","UI/UX","Splash Screen Loads & Displays Branding"),
            ("TC-WEB-002","Splash","Functional","Splash Auto-Navigation (No Session → Login/Onboarding)"),
            ("TC-WEB-003","Onboarding","UI/UX","Onboarding Slide 1 — Get Started Navigates to Slide 2"),
            ("TC-WEB-004","Onboarding","UI/UX","Onboarding Slide 2 — Continue Navigates to Slide 3"),
            ("TC-WEB-005","Onboarding","Functional","Onboarding Skip Button → Navigates to Login Page"),
            ("TC-WEB-006","Authentication","UI/UX","Login Page Loads — Email, Password & Sign In Fields Present"),
            ("TC-WEB-007","Authentication","Validation","Login with Valid Credentials → Navigates to Dashboard"),
            ("TC-WEB-008","Authentication","Validation","Login with Wrong Password — Error Snackbar Shown"),
            ("TC-WEB-009","Authentication","Validation","Login with Empty Email — Validation Error Displayed"),
            ("TC-WEB-010","Authentication","UI/UX","Password Visibility Toggle (Show/Hide Eye Icon)"),
            ("TC-WEB-011","Authentication","Functional","Toggle Login Page: Sign In Mode → Sign Up Mode"),
            ("TC-WEB-012","Authentication","Validation","Register with Mismatched Passwords — Validation Error"),
            ("TC-WEB-013","Authentication","Validation","Register with Short Password (<6 chars) — Validation Error"),
            ("TC-WEB-014","Authentication","Functional","Forgot Password Link → Navigates to Password Reset Page"),
            ("TC-WEB-015","Authentication","Validation","Forgot Password — Invalid Email Format Blocked by Validation"),
            ("TC-WEB-016","Authentication","Functional","Forgot Password — Valid Email Triggers OTP Send & OTP Form Appears"),
            ("TC-WEB-017","Authentication","UI/UX","Forgot Password — 5-Minute OTP Expiry Countdown Timer Displayed"),
            ("TC-WEB-018","Authentication","Validation","Suggest Password Feature — Auto-fills New Password Field"),
            ("TC-WEB-019","Authentication","Functional","Forgot Password — Change Email Returns to Email Entry Form"),
            ("TC-WEB-020","Authentication","Validation","Full Login Flow — Session Created & Dashboard Confirmed"),
            ("TC-WEB-021","Navigation","Functional","Bottom Navigation — Analytics Tab Loads Analytics Page"),
            ("TC-WEB-022","Navigation","Functional","Bottom Navigation — History Tab Loads Patient History Screen"),
            ("TC-WEB-023","Navigation","Functional","Bottom Navigation — Settings Tab Loads Settings Page"),
            ("TC-WEB-024","Navigation","Functional","Bottom Navigation — Profile Tab Loads Profile Page"),
            ("TC-WEB-025","Navigation","Functional","FAB (+ Button) Opens New Case Page"),
            ("TC-WEB-026","Navigation","UI/UX","App Branding — Title/Logo Visible on Authenticated Pages"),
            ("TC-WEB-027","Dashboard","UI/UX","Dashboard Loads — Greeting, AI Status Badge & Stats Visible"),
            ("TC-WEB-028","Dashboard","Functional","Notification Bell — Clinical Alerts Bottom Sheet Opens"),
            ("TC-WEB-029","Dashboard","Functional","Doctor Avatar Tap → Navigates to Profile Page"),
            ("TC-MOB-030","Dashboard","Functional","New Case Assessment CTA → Navigates to New Case Page"),
            ("TC-WEB-031","Dashboard","Functional","Recent Assessment Card Tap → Navigates to Case Detail Page"),
            ("TC-WEB-032","New Case","UI/UX","New Case Page Loads — Patient ID Field & 4-Section Layout Visible"),
            ("TC-WEB-033","New Case","Functional","Patient ID Search — Existing Patient Auto-fills All Fields"),
            ("TC-WEB-034","New Case","UI/UX","Patient ID Search — Unknown Patient Shows 'New Patient' Snackbar"),
            ("TC-WEB-035","New Case","Validation","Proceed Without Required Fields — Validation Snackbar Shown"),
            ("TC-WEB-036","New Case","UI/UX","Patient Photo Area — Camera Tap Element Accessible"),
            ("TC-WEB-037","New Case","Functional","Sex Dropdown Opens — Male/Female/Other Options Accessible"),
            ("TC-WEB-038","New Case","Functional","Complete Patient Form + Proceed → Navigates to Image Upload"),
            ("TC-WEB-039","New Case (Sect A)","Functional","Section A — Smoking Habit Choice Buttons (No/Past/Current) Clickable"),
            ("TC-WEB-040","New Case (Sect A)","Functional","Section A — Smokeless Tobacco Switch Toggle Works"),
            ("TC-WEB-041","New Case (Sect A)","Functional","Section A — Diabetes Medical Condition Switch Toggle"),
            ("TC-WEB-042","New Case (Sect B)","Functional","Section B — Lesion Duration Choice Pills Clickable"),
            ("TC-WEB-043","New Case (Sect B)","Functional","Section B — Recurrence Pattern Dropdown Opens"),
            ("TC-WEB-044","New Case (Sect C)","Functional","Section C — Anatomical Site Dropdown Opens (Risk Sites Shown)"),
            ("TC-WEB-045","New Case (Sect C)","Functional","Section C — Lesion Size Stepper (+/-) Increments & Decrements"),
            ("TC-WEB-046","New Case (Sect C)","Functional","Section C — Induration Present Switch Toggle"),
            ("TC-WEB-047","New Case (Sect D)","Functional","Section D — Palpable Lymph Node Switch Toggle"),
            ("TC-WEB-048","Image Upload","UI/UX","Image Upload Page — Camera & Gallery Buttons Present"),
            ("TC-WEB-049","Image Upload","Validation","Process AI Button — Disabled State When No Image Selected"),
            ("TC-WEB-050","Image Upload","UI/UX","Photography Tips Card — Guidance Chips (Lighting, Focus, Shadows) Visible"),
            ("TC-WEB-051","Image Upload","Functional","Gallery Button Tap — File Picker Accessible on Desktop"),
            ("TC-WEB-052","Image Upload","Functional","Back Button on Image Upload — Returns to Previous Page"),
            ("TC-WEB-053","AI Result","UI/UX","AI Result Page Loads — Risk Banner & Score Gauges Visible"),
            ("TC-WEB-054","AI Result","UI/UX","AI Confidence Percentage Score — Displayed on Result Screen"),
            ("TC-WEB-055","AI Result","UI/UX","Key Contributing Risk Factors — Listed on AI Result Screen"),
            ("TC-WEB-056","AI Result","Functional","Return to Dashboard Button — Navigates Back to Dashboard"),
            ("TC-WEB-057","History","UI/UX","History Screen Loads — Patient Case List / Grid Visible"),
            ("TC-WEB-058","History","Functional","Search Bar — Real-time Filter by Patient Name / ID / Doctor"),
            ("TC-WEB-059","History","Functional","Search Clear (X) Button — Resets Search Query & Shows All Cases"),
            ("TC-WEB-060","History","Functional","Risk Filter Chip — HIGH Risk Filters Case List Correctly"),
            ("TC-WEB-061","History","Functional","Risk Filter Chip — INTERMEDIATE Risk Filters Case List"),
            ("TC-WEB-062","History","Functional","Clinical Filter Chip — Biopsy Recommendation Filter Toggle"),
            ("TC-WEB-063","History","Functional","Toggle View — Switch Between List View and Grid (Gallery) View"),
            ("TC-WEB-064","History","Functional","Deduplicate Toggle — Switch Between All Visits & Unique Patients View"),
            ("TC-WEB-065","History","Functional","Date Range Filter — Date Picker Dialog Opens"),
            ("TC-WEB-066","History","Functional","Quick Actions Sheet — Opens on Case Card Tap (View/Edit/Export Options)"),
            ("TC-WEB-067","History","Functional","Long Press — Enters Multi-Select Mode (Select All Button Visible)"),
            ("TC-WEB-068","History","Functional","Export CSV — Current Filtered Case List Exported as CSV File"),
            ("TC-WEB-069","Case Detail","UI/UX","Case Detail Page Loads — Patient Strip, Risk Banner & Sections A-D Visible"),
            ("TC-WEB-070","Case Detail","UI/UX","Risk Progression Line Chart — Visible for Multi-Visit Patients"),
            ("TC-WEB-071","Case Detail","UI/UX","Section A — Patient Demographics Data Table Visible"),
            ("TC-WEB-072","Case Detail","Functional","Generate Clinical PDF Report — Share Button Triggers PDF Creation"),
            ("TC-WEB-073","Analytics","UI/UX","Analytics Page Loads — Stats Boxes, Charts & Filters Visible"),
            ("TC-WEB-074","Analytics","Functional","Time Range Selector — 7D Option Filters All Charts to Last 7 Days"),
            ("TC-WEB-075","Analytics","Functional","Time Range Selector — ALL Option Resets to All-Time Analytics"),
            ("TC-WEB-076","Analytics","Functional","Analytics Search Bar — Patient Name / ID Query Filters All Metrics"),
            ("TC-WEB-077","Analytics","Functional","Filter Chip — Indurated Toggle Filters Cases with Induration"),
            ("TC-WEB-078","Analytics","Functional","Anatomical Heatmap — Tongue Tile Tap Filters All Metrics to That Site"),
            ("TC-WEB-079","Analytics","Functional","High Risk Stat Box — Tap Filters Dashboard to HIGH Risk Cases Only"),
            ("TC-WEB-080","Analytics","Functional","Clear All Filters Button — Resets All Filters"),
            ("TC-WEB-081","Analytics","Functional","Export PDF Audit Report — Print Button Generates A4 Analytics PDF"),
            ("TC-WEB-082","Profile","UI/UX","Profile Page Loads — Doctor Name, Email, Department & Milestones Visible"),
            ("TC-WEB-083","Profile","Functional","Profile Photo — Tap Opens Camera/Gallery Bottom Sheet"),
            ("TC-WEB-084","Profile","Functional","Achievement Badge — Screening Rookie Tap Opens Achievement Detail Dialog"),
            ("TC-WEB-085","Profile","UI/UX","Verified Badge — License Dialog Opens with DCI Registration Field"),
            ("TC-WEB-086","Profile","Functional","Digital Signature — Canvas Dialog Opens with Clear & Save Buttons"),
            ("TC-WEB-087","Profile","Functional","Print/Save ID Badge — CR80-Format PDF ID Badge Triggers Successfully"),
            ("TC-WEB-088","Settings","UI/UX","Settings Page Loads — All 7 Sections Present"),
            ("TC-WEB-089","Settings","Functional","Preferences — High-Risk Alert Banner Switch Toggle Saves Setting"),
            ("TC-WEB-090","Settings","Functional","Preferences — Compact Case List Switch Toggle Saves Setting"),
            ("TC-WEB-091","Settings","Functional","Department Dropdown — Opens & Allows Selection of Active Department"),
            ("TC-WEB-092","Settings","Functional","PDF Settings — Include Signature Line Switch Toggle"),
            ("TC-WEB-093","Settings","Functional","Privacy — Require Digital Consent Switch"),
            ("TC-WEB-094","Settings","Functional","Clear Offline Cache — Tile Tap Clears Cache & Shows Success Snackbar"),
            ("TC-WEB-095","Settings","Functional","Export Database to CSV — Exports All Cases as CSV & Shows Dialog"),
            ("TC-WEB-096","Settings","Functional","Monthly Audit PDF — Generates A4 Audit Report & Opens Print Dialog"),
            ("TC-WEB-097","Settings","UI/UX","Server Settings Dialog — Opens with URL Field & Reset/Save Buttons"),
            ("TC-WEB-098","Settings","Functional","Server Dialog — Reset to Default Button Restores HF Space URL"),
            ("TC-WEB-099","Settings","Validation","Server Dialog — Save & Test Saves URL & Triggers Live Connection Ping"),
            ("TC-WEB-100","Settings","Validation","Live Connection Ping — Tap Triggers API Ping & Shows Latency/Offline"),
            ("TC-WEB-101","Settings","UI/UX","How AI Works — Dialog Opens Showing Consensus Engine Explanation"),
            ("TC-WEB-102","Settings","UI/UX","Saveetha Diagnostic Protocol — 6-Step Handbook Dialog Opens"),
            ("TC-WEB-103","Settings","UI/UX","Medical Disclaimer — 5-Item Disclaimer Dialog Opens"),
            ("TC-WEB-104","Settings","Functional","Sign Out (Settings) — Clears Session & Redirects to Login"),
            ("TC-WEB-105","Profile","Functional","Sign Out (Profile) — Clears Session & Redirects to Login Page"),
        ]

        mock_results = []
        for tc_id, module, test_type, desc in MOCK_TESTS:
            status   = "PASS" if random.random() > 0.12 else "FAIL"
            duration = random.randint(450, 5500)
            log_msg  = "No errors. Flow completed successfully." if status == "PASS" else "Element not found within timeout."
            mock_results.append({
                "id": tc_id, "platform": "Desktop Web", "test_type": test_type,
                "module": module, "description": desc,
                "status": status, "duration": duration, "logs": log_msg
            })

        generate_report(mock_results, "Web_Test_Execution_Report.xlsx")
