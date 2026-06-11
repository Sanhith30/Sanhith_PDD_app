"""
report_generator.py — Excel Report Generator
Oral Ulcer AI — Saveetha Dental College
Surgical Luxury Theme (Maroon #7B1E3A / Gold #C9A84C)
"""
import os
import sys
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, GradientFill
from openpyxl.utils import get_column_letter


def generate_report(results, filename="Web_Test_Execution_Report.xlsx"):
    """
    Generates a beautifully styled Excel spreadsheet containing test run results.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Web Test Execution"

    # ── Color Palette ─────────────────────────────────────────────
    MAROON      = "7B1E3A"
    DARK_MAROON = "5C1028"
    GOLD        = "C9A84C"
    IVORY       = "FAF7F4"
    PASS_BG     = "E8F5E9"
    FAIL_BG     = "FFEBEE"
    PASS_FG     = "1B5E20"
    FAIL_FG     = "B71C1C"
    STRIPE_BG   = "F5F0ED"

    # ── Styles ────────────────────────────────────────────────────
    maroon_fill  = PatternFill("solid", fgColor=MAROON)
    stripe_fill  = PatternFill("solid", fgColor=STRIPE_BG)
    ivory_fill   = PatternFill("solid", fgColor=IVORY)
    pass_fill    = PatternFill("solid", fgColor=PASS_BG)
    fail_fill    = PatternFill("solid", fgColor=FAIL_BG)

    header_font  = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    title_font   = Font(name="Calibri", size=14, bold=True, color=MAROON)
    subtitle_font= Font(name="Calibri", size=9,  italic=True, color="9E8A8F")
    regular_font = Font(name="Calibri", size=9)
    pass_font    = Font(name="Calibri", size=9, bold=True, color=PASS_FG)
    fail_font    = Font(name="Calibri", size=9, bold=True, color=FAIL_FG)
    id_font      = Font(name="Calibri", size=9, bold=True, color=MAROON)
    gold_font    = Font(name="Calibri", size=9, bold=True, color="A07828")
    mod_font     = Font(name="Calibri", size=9, bold=True, color="4A4A4A")

    thin = Side(style="thin",   color="D8C8C0")
    med  = Side(style="medium", color=GOLD)
    data_border  = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_border= Border(
        left=Side(style="thin", color=DARK_MAROON),
        right=Side(style="thin", color=DARK_MAROON),
        top=Side(style="thin", color=DARK_MAROON),
        bottom=Side(style="medium", color=GOLD)
    )

    center  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ── Title Banner (Rows 1–3) ────────────────────────────────────
    ws.merge_cells("A1:G1")
    ws["A1"] = "ORAL ULCER AI — WEB APPLICATION E2E TEST REPORT"
    ws["A1"].font      = title_font
    ws["A1"].alignment = center
    ws["A1"].fill      = PatternFill("solid", fgColor="FAF7F4")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    now = datetime.datetime.now().strftime("%d %B %Y  |  %H:%M:%S")
    ws["A2"] = f"Saveetha Dental College & Hospital  •  Generated: {now}"
    ws["A2"].font      = subtitle_font
    ws["A2"].alignment = center
    ws.row_dimensions[2].height = 18

    ws.merge_cells("A3:G3")
    ws["A3"] = ""
    ws.row_dimensions[3].height = 6

    # ── Summary Stats Row (Row 4) ──────────────────────────────────
    total   = len(results)
    passed  = sum(1 for r in results if r.get("status") == "PASS")
    failed  = total - passed
    rate    = f"{round(passed / total * 100)}%" if total else "0%"
    modules = len(set(r.get("module", "") for r in results))

    stats = [
        ("TOTAL TCs", str(total),  MAROON),
        ("PASSED",    str(passed), "2E7D32"),
        ("FAILED",    str(failed), "C62828"),
        ("PASS RATE", rate,        "A07828"),
        ("MODULES",   str(modules), "37474F"),
    ]
    col_positions = [1, 2, 3, 4, 5]
    for (label, value, color), col in zip(stats, col_positions):
        label_cell = ws.cell(row=4, column=col)
        label_cell.value     = label
        label_cell.font      = Font(name="Calibri", size=8, bold=True, color="888888")
        label_cell.alignment = center
        label_cell.fill      = PatternFill("solid", fgColor="F0EAE6")

        val_cell = ws.cell(row=5, column=col)
        val_cell.value     = value
        val_cell.font      = Font(name="Calibri", size=18, bold=True, color=color)
        val_cell.alignment = center
        val_cell.fill      = PatternFill("solid", fgColor="F0EAE6")

    ws.row_dimensions[4].height = 16
    ws.row_dimensions[5].height = 34

    ws.merge_cells("F4:G4")
    ws.merge_cells("F5:G5")
    ws["F4"].value     = "AUTOMATED E2E TESTING"
    ws["F4"].font      = Font(name="Calibri", size=8, bold=True, color="888888")
    ws["F4"].alignment = center
    ws["F5"].value     = "Selenium WebDriver / Appium Chrome"
    ws["F5"].font      = Font(name="Calibri", size=9, color="4A4A4A", italic=True)
    ws["F5"].alignment = center

    ws.row_dimensions[6].height = 8

    # ── Column Headers (Row 7) ────────────────────────────────────
    headers = [
        "Test Case ID", "Platform", "Module",
        "Test Description", "Status", "Duration (ms)",
        "Error / Output Logs"
    ]
    ws.append([""] * 7)  # row 7 placeholder
    ws.append(headers)   # row 8

    header_row_idx = 8
    ws.row_dimensions[header_row_idx].height = 28
    for col, cell in enumerate(ws[header_row_idx], 1):
        cell.font      = header_font
        cell.fill      = maroon_fill
        cell.alignment = center
        cell.border    = header_border

    # ── Data Rows ─────────────────────────────────────────────────
    prev_module = None
    for i, test in enumerate(results):
        row_data = [
            test.get("id", ""),
            test.get("platform", "Desktop Web"),
            test.get("module", ""),
            test.get("description", ""),
            test.get("status", ""),
            test.get("duration", 0),
            test.get("logs", "No errors. Flow completed successfully."),
        ]
        ws.append(row_data)
        row_idx    = ws.max_row
        is_stripe  = (i % 2 == 1)
        row_fill   = stripe_fill if is_stripe else ivory_fill
        status     = test.get("status", "")
        cur_module = test.get("module", "")

        ws.row_dimensions[row_idx].height = 22

        for col_idx, cell in enumerate(ws[row_idx], 1):
            cell.border    = data_border
            cell.alignment = left

            if col_idx == 1:  # ID
                cell.font      = id_font
                cell.alignment = center
                cell.fill      = row_fill
            elif col_idx == 2:  # Platform
                cell.font      = Font(name="Calibri", size=8, color="6A5560")
                cell.alignment = center
                cell.fill      = row_fill
            elif col_idx == 3:  # Module
                cell.font  = mod_font
                cell.fill  = row_fill
                if cur_module != prev_module:
                    cell.font = Font(name="Calibri", size=9, bold=True, color=MAROON)
                prev_module = cur_module
            elif col_idx == 4:  # Description
                cell.font = regular_font
                cell.fill = row_fill
            elif col_idx == 5:  # Status
                cell.alignment = center
                if status == "PASS":
                    cell.font = pass_font
                    cell.fill = pass_fill
                else:
                    cell.font = fail_font
                    cell.fill = fail_fill
            elif col_idx == 6:  # Duration
                cell.alignment = center
                cell.font      = Font(name="Calibri", size=9, color="4A4A4A")
                cell.fill      = row_fill
            elif col_idx == 7:  # Logs
                cell.font = Font(name="Calibri", size=8, color="6A5560", italic=(status == "PASS"))
                cell.fill = row_fill

    # ── Column Widths ─────────────────────────────────────────────
    widths = {"A": 14, "B": 18, "C": 22, "D": 52, "E": 11, "F": 14, "G": 55}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    # ── Freeze Panes ──────────────────────────────────────────────
    ws.freeze_panes = "A9"

    # ── Save ──────────────────────────────────────────────────────
    reports_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "reports")
    os.makedirs(reports_dir, exist_ok=True)
    output_path = os.path.join(reports_dir, filename)
    wb.save(output_path)

    print("\n" + "=" * 58)
    print("  [SUCCESS] Test report generated successfully!")
    print(f"  Path   : {output_path}")
    print(f"  Total  : {total}  |  Passed: {passed}  |  Failed: {failed}  |  Rate: {rate}")
    print("=" * 58 + "\n")


# ── Dry-Run (mock data for 105 tests) ─────────────────────────────
if __name__ == "__main__":
    if "--dry" in sys.argv:
        import random
        print("Generating comprehensive mock test report (105 test cases)...")

        MOCK_TESTS = [
            # Splash & Onboarding
            ("TC-WEB-001","Splash","Splash Screen Loads & Displays Branding"),
            ("TC-WEB-002","Splash","Splash Auto-Navigation (No Session → Login/Onboarding)"),
            ("TC-WEB-003","Onboarding","Onboarding Slide 1 — Get Started Navigates to Slide 2"),
            ("TC-WEB-004","Onboarding","Onboarding Slide 2 — Continue Navigates to Slide 3"),
            ("TC-WEB-005","Onboarding","Onboarding Skip Button → Navigates to Login Page"),
            # Authentication
            ("TC-WEB-006","Authentication","Login Page Loads — Email, Password & Sign In Fields Present"),
            ("TC-WEB-007","Authentication","Login with Valid Credentials → Navigates to Dashboard"),
            ("TC-WEB-008","Authentication","Login with Wrong Password — Error Snackbar Shown"),
            ("TC-WEB-009","Authentication","Login with Empty Email — Validation Error Displayed"),
            ("TC-WEB-010","Authentication","Password Visibility Toggle (Show/Hide Eye Icon)"),
            ("TC-WEB-011","Authentication","Toggle Login Page: Sign In Mode → Sign Up Mode"),
            ("TC-WEB-012","Authentication","Register with Mismatched Passwords — Validation Error"),
            ("TC-WEB-013","Authentication","Register with Short Password (<6 chars) — Validation Error"),
            ("TC-WEB-014","Authentication","Forgot Password Link → Navigates to Password Reset Page"),
            ("TC-WEB-015","Authentication","Forgot Password — Invalid Email Format Blocked by Validation"),
            ("TC-WEB-016","Authentication","Forgot Password — Valid Email Triggers OTP Send & OTP Form Appears"),
            ("TC-WEB-017","Authentication","Forgot Password — 5-Minute OTP Expiry Countdown Timer Displayed"),
            ("TC-WEB-018","Authentication","Suggest Password Feature — Auto-fills New Password Field"),
            ("TC-WEB-019","Authentication","Forgot Password — Change Email Returns to Email Entry Form"),
            ("TC-WEB-020","Authentication","Full Login Flow — Session Created & Dashboard Confirmed"),
            # Navigation
            ("TC-WEB-021","Navigation","Bottom Navigation — Analytics Tab Loads Analytics Page"),
            ("TC-WEB-022","Navigation","Bottom Navigation — History Tab Loads Patient History Screen"),
            ("TC-WEB-023","Navigation","Bottom Navigation — Settings Tab Loads Settings Page"),
            ("TC-WEB-024","Navigation","Bottom Navigation — Profile Tab Loads Profile Page"),
            ("TC-WEB-025","Navigation","FAB (+ Button) Opens New Case Page"),
            ("TC-WEB-026","Navigation","App Branding — Title/Logo Visible on Authenticated Pages"),
            # Dashboard
            ("TC-WEB-027","Dashboard","Dashboard Loads — Greeting, AI Status Badge & Stats Visible"),
            ("TC-WEB-028","Dashboard","Notification Bell — Clinical Alerts Bottom Sheet Opens"),
            ("TC-WEB-029","Dashboard","Doctor Avatar Tap → Navigates to Profile Page"),
            ("TC-WEB-030","Dashboard","New Case Assessment CTA → Navigates to New Case Page"),
            ("TC-WEB-031","Dashboard","Recent Assessment Card Tap → Navigates to Case Detail Page"),
            # New Case — Patient Identification
            ("TC-WEB-032","New Case","New Case Page Loads — Patient ID Field & 4-Section Layout Visible"),
            ("TC-WEB-033","New Case","Patient ID Search — Existing Patient Auto-fills All Fields"),
            ("TC-WEB-034","New Case","Patient ID Search — Unknown Patient Shows 'New Patient' Snackbar"),
            ("TC-WEB-035","New Case","Proceed Without Required Fields — Validation Snackbar Shown"),
            ("TC-WEB-036","New Case","Patient Photo Area — Camera Tap Element Accessible"),
            ("TC-WEB-037","New Case","Sex Dropdown Opens — Male/Female/Other Options Accessible"),
            ("TC-WEB-038","New Case","Complete Patient Form + Proceed → Navigates to Image Upload"),
            # New Case — Sections A–D
            ("TC-WEB-039","New Case (Sect A)","Section A — Smoking Habit Choice Buttons (No/Past/Current) Clickable"),
            ("TC-WEB-040","New Case (Sect A)","Section A — Smokeless Tobacco Switch Toggle Works"),
            ("TC-WEB-041","New Case (Sect A)","Section A — Diabetes Medical Condition Switch Toggle"),
            ("TC-WEB-042","New Case (Sect B)","Section B — Lesion Duration Choice Pills Clickable"),
            ("TC-WEB-043","New Case (Sect B)","Section B — Recurrence Pattern Dropdown Opens"),
            ("TC-WEB-044","New Case (Sect C)","Section C — Anatomical Site Dropdown Opens (Risk Sites Shown)"),
            ("TC-WEB-045","New Case (Sect C)","Section C — Lesion Size Stepper (+/-) Increments & Decrements"),
            ("TC-WEB-046","New Case (Sect C)","Section C — Induration Present Switch Toggle"),
            ("TC-WEB-047","New Case (Sect D)","Section D — Palpable Lymph Node Switch Toggle"),
            # Image Upload
            ("TC-WEB-048","Image Upload","Image Upload Page — Camera & Gallery Buttons Present"),
            ("TC-WEB-049","Image Upload","Process AI Button — Disabled State When No Image Selected"),
            ("TC-WEB-050","Image Upload","Photography Tips Card — Guidance Chips (Lighting, Focus, Shadows) Visible"),
            ("TC-WEB-051","Image Upload","Gallery Button Tap — File Picker Accessible on Web"),
            ("TC-WEB-052","Image Upload","Back Button on Image Upload — Returns to Previous Page"),
            # AI Result
            ("TC-WEB-053","AI Result","AI Result Page Loads — Risk Banner & Score Gauges Visible"),
            ("TC-WEB-054","AI Result","AI Confidence Percentage Score — Displayed on Result Screen"),
            ("TC-WEB-055","AI Result","Key Contributing Risk Factors — Listed on AI Result Screen"),
            ("TC-WEB-056","AI Result","Return to Dashboard Button — Navigates Back to Dashboard"),
            # History Screen
            ("TC-WEB-057","History","History Screen Loads — Patient Case List / Grid Visible"),
            ("TC-WEB-058","History","Search Bar — Real-time Filter by Patient Name / ID / Doctor"),
            ("TC-WEB-059","History","Search Clear (X) Button — Resets Search Query & Shows All Cases"),
            ("TC-WEB-060","History","Risk Filter Chip — HIGH Risk Filters Case List Correctly"),
            ("TC-WEB-061","History","Risk Filter Chip — INTERMEDIATE Risk Filters Case List"),
            ("TC-WEB-062","History","Clinical Filter Chip — Biopsy Recommendation Filter Toggle"),
            ("TC-WEB-063","History","Toggle View — Switch Between List View and Grid (Gallery) View"),
            ("TC-WEB-064","History","Deduplicate Toggle — Switch Between All Visits & Unique Patients View"),
            ("TC-WEB-065","History","Date Range Filter — Date Picker Dialog Opens"),
            ("TC-WEB-066","History","Quick Actions Sheet — Opens on Case Card Tap (View/Edit/Export Options)"),
            ("TC-WEB-067","History","Long Press — Enters Multi-Select Mode (Select All Button Visible)"),
            ("TC-WEB-068","History","Export CSV — Current Filtered Case List Exported as CSV File"),
            # Case Detail
            ("TC-WEB-069","Case Detail","Case Detail Page Loads — Patient Strip, Risk Banner & Sections A-D Visible"),
            ("TC-WEB-070","Case Detail","Risk Progression Line Chart — Visible for Multi-Visit Patients"),
            ("TC-WEB-071","Case Detail","Section A — Patient Demographics Data Table Visible"),
            ("TC-WEB-072","Case Detail","Generate Clinical PDF Report — Share Button Triggers PDF Creation"),
            # Analytics
            ("TC-WEB-073","Analytics","Analytics Page Loads — Stats Boxes, Charts & Filters Visible"),
            ("TC-WEB-074","Analytics","Time Range Selector — 7D Option Filters All Charts to Last 7 Days"),
            ("TC-WEB-075","Analytics","Time Range Selector — ALL Option Resets to All-Time Analytics"),
            ("TC-WEB-076","Analytics","Analytics Search Bar — Patient Name / ID Query Filters All Metrics"),
            ("TC-WEB-077","Analytics","Filter Chip — Indurated Toggle Filters Cases with Induration"),
            ("TC-WEB-078","Analytics","Anatomical Heatmap — Tongue Tile Tap Filters All Metrics to That Site"),
            ("TC-WEB-079","Analytics","High Risk Stat Box — Tap Filters Dashboard to HIGH Risk Cases Only"),
            ("TC-WEB-080","Analytics","Clear All Filters Button — Resets All Filters"),
            ("TC-WEB-081","Analytics","Export PDF Audit Report — Print Button Generates A4 Analytics PDF"),
            # Profile
            ("TC-WEB-082","Profile","Profile Page Loads — Doctor Name, Email, Department & Milestones Visible"),
            ("TC-WEB-083","Profile","Profile Photo — Tap Opens Camera/Gallery Bottom Sheet"),
            ("TC-WEB-084","Profile","Achievement Badge — Screening Rookie Tap Opens Achievement Detail Dialog"),
            ("TC-WEB-085","Profile","Verified Badge — License Dialog Opens with DCI Registration Field"),
            ("TC-WEB-086","Profile","Digital Signature — Canvas Dialog Opens with Clear & Save Buttons"),
            ("TC-WEB-087","Profile","Print/Save ID Badge — CR80-Format PDF ID Badge Triggers Successfully"),
            # Settings
            ("TC-WEB-088","Settings","Settings Page Loads — All 7 Sections Present"),
            ("TC-WEB-089","Settings","Preferences — High-Risk Alert Banner Switch Toggle Saves Setting"),
            ("TC-WEB-090","Settings","Preferences — Compact Case List Switch Toggle Saves Setting"),
            ("TC-WEB-091","Settings","Department Dropdown — Opens & Allows Selection of Active Department"),
            ("TC-WEB-092","Settings","PDF Settings — Include Signature Line Switch Toggle"),
            ("TC-WEB-093","Settings","Privacy — Require Digital Consent Switch"),
            ("TC-WEB-094","Settings","Clear Offline Cache — Tile Tap Clears Cache & Shows Success Snackbar"),
            ("TC-WEB-095","Settings","Export Database to CSV — Exports All Cases as CSV & Shows Dialog"),
            ("TC-WEB-096","Settings","Monthly Audit PDF — Generates A4 Audit Report & Opens Print Dialog"),
            ("TC-WEB-097","Settings","Server Settings Dialog — Opens with URL Field & Reset/Save Buttons"),
            ("TC-WEB-098","Settings","Server Dialog — Reset to Default Button Restores HF Space URL"),
            ("TC-WEB-099","Settings","Server Dialog — Save & Test Saves URL & Triggers Live Connection Ping"),
            ("TC-WEB-100","Settings","Live Connection Ping — Tap Triggers API Ping & Shows Latency/Offline"),
            ("TC-WEB-101","Settings","How AI Works — Dialog Opens Showing Consensus Engine Explanation"),
            ("TC-WEB-102","Settings","Saveetha Diagnostic Protocol — 6-Step Handbook Dialog Opens"),
            ("TC-WEB-103","Settings","Medical Disclaimer — 5-Item Disclaimer Dialog Opens"),
            ("TC-WEB-104","Settings","Sign Out (Settings) — Clears Session & Redirects to Login"),
            ("TC-WEB-105","Profile","Sign Out (Profile) — Clears Session & Redirects to Login Page"),
        ]

        mock_results = []
        for tc_id, module, desc in MOCK_TESTS:
            status   = "PASS" if random.random() > 0.12 else "FAIL"
            duration = random.randint(450, 5500)
            log_msg  = "No errors. Flow completed successfully." if status == "PASS" else "Element not found within timeout. Check ARIA labels on Flutter Web semantic tree."
            mock_results.append({
                "id": tc_id, "platform": "Desktop Web",
                "module": module, "description": desc,
                "status": status, "duration": duration, "logs": log_msg
            })

        generate_report(mock_results, "Web_Test_Execution_Report.xlsx")
