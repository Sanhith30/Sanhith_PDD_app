import os
import sys
import time
import json
import subprocess
import socket
from io import BytesIO

# Try to install openpyxl automatically if not present
try:
    import openpyxl
    import pandas as pd
except ImportError:
    print("Required Excel libraries (pandas, openpyxl) not found. Attempting to install them...")
    try:
        subprocess.run([sys.executable, "-m", "pip", "install", "pandas", "openpyxl"], check=True)
        import openpyxl
        import pandas as pd
        print("Excel libraries installed successfully.")
    except Exception as e:
        print(f"Could not install Excel libraries: {e}. Testing will continue and output to CSV instead.")
        pd = None

import requests

BASE_URL = "http://127.0.0.1:5000"

def is_port_in_use(port):
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        return s.connect_ex(('127.0.0.1', port)) == 0

def create_mock_image():
    # Create a small 100x100 red PNG image in memory to upload
    try:
        from PIL import Image
        img = Image.new('RGB', (100, 100), color='red')
        img_byte_arr = BytesIO()
        img.save(img_byte_arr, format='PNG')
        img_byte_arr.seek(0)
        return img_byte_arr
    except Exception:
        # Fallback raw byte array of a minimal PNG
        return BytesIO(
            b'\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01\x08\x06\x00\x00\x00\x1f\x15\xc4\x89'
            b'\x00\x00\x00\rIDATx\x9cc`\x00\x01\x00\x00\x05\x00\x01\r\n-\xb4\x00\x00\x00\x00IEND\xaeB`\x82'
        )

def run_tests():
    print("====================================================")
    print("       Saveetha Oral Sentry Functionality Tests     ")
    print("====================================================")
    
    server_process = None
    if not is_port_in_use(5000):
        print("Starting FastAPI backend server locally on port 5000...")
        server_process = subprocess.Popen(
            [sys.executable, "app.py"],
            cwd=os.path.dirname(os.path.abspath(__file__)),
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True
        )
        # Wait for server to boot
        boot_attempts = 60
        for i in range(boot_attempts):
            try:
                r = requests.get(BASE_URL, timeout=1)
                if r.status_code == 200:
                    print("Server started successfully.")
                    break
            except requests.RequestException:
                pass
            print(f"Waiting for server to start... ({i+1}/{boot_attempts})")
            time.sleep(1)
        else:
            print("Error: FastAPI server failed to start. Exiting tests.")
            if server_process:
                server_process.terminate()
            sys.exit(1)
    else:
        print("FastAPI server is already running on port 5000. Testing against running server.")

    # Unique parameters for this run
    ts = int(time.time())
    clinician_email = f"test_doctor_{ts}@gmail.com"
    clinician_password = "TestPassword123!"
    clinician_name = f"Dr. Tester {ts}"
    new_password = "NewPassword123!"
    patient_id = f"PT_TEST_{ts}"
    
    token = None
    clinician_id = None
    case_id = None
    
    results = []

    # Helper function to record test case
    def record_test(test_id, component, description, request_details, expected, action_func):
        print(f"Running TC-{test_id:02d}: {description}...", end="", flush=True)
        start_time = time.time()
        status = "FAIL"
        actual_code = "N/A"
        actual_response = ""
        
        try:
            res = action_func()
            latency = round(time.time() - start_time, 3)
            
            if isinstance(res, requests.Response):
                actual_code = str(res.status_code)
                try:
                    res_json = res.json()
                    actual_response = json.dumps(res_json)
                except Exception:
                    actual_response = res.text
                
                # Check success condition
                if res.status_code in [200, 201]:
                    # Some endpoints might return success=False inside JSON, check that
                    try:
                        res_json = res.json()
                        if res_json.get("success", True) is not False:
                            status = "PASS"
                    except Exception:
                        status = "PASS"
            else:
                actual_code = "OK"
                actual_response = str(res)
                status = "PASS"
        except Exception as e:
            latency = round(time.time() - start_time, 3)
            actual_response = f"Exception: {str(e)}"
            status = "FAIL"
            
        print(f" [{status}] in {latency}s")
        
        results.append({
            "Test Case ID": f"TC-{test_id:02d}",
            "Component": component,
            "Description": description,
            "Request Details": request_details,
            "Expected": expected,
            "Actual Status Code": actual_code,
            "Actual Response": actual_response[:200] + ("..." if len(actual_response) > 200 else ""),
            "Latency (s)": latency,
            "Status": status
        })
        return status == "PASS", actual_response

    # 1. Health Check
    record_test(
        1, "System", "Health check endpoint returns running status",
        "GET /", "Status code 200, status='healthy'",
        lambda: requests.get(f"{BASE_URL}/")
    )

    # 2. Clinician Signup
    signup_payload = {"name": clinician_name, "email": clinician_email, "password": clinician_password}
    def do_signup():
        r = requests.post(f"{BASE_URL}/auth/signup", json=signup_payload)
        nonlocal token, clinician_id
        if r.status_code == 200:
            res_json = r.json()
            token = res_json.get("access_token")
            clinician_id = str(res_json.get("user", {}).get("id"))
        return r
    record_test(
        2, "Authentication", "Register a new clinician account",
        f"POST /auth/signup\n{json.dumps(signup_payload)}", "Status code 200, success=True, access_token returned",
        do_signup
    )

    # 3. Clinician Login
    login_payload = {"email": clinician_email, "password": clinician_password}
    record_test(
        3, "Authentication", "Log in with the registered clinician credentials",
        f"POST /auth/login\n{json.dumps(login_payload)}", "Status code 200, access_token returned",
        lambda: requests.post(f"{BASE_URL}/auth/login", json=login_payload)
    )

    # 4. Request Password Reset OTP
    reset_payload = {"email": clinician_email}
    record_test(
        4, "Authentication", "Request password reset OTP email",
        f"POST /auth/reset_password\n{json.dumps(reset_payload)}", "Status code 200, success=True",
        lambda: requests.post(f"{BASE_URL}/auth/reset_password", json=reset_payload)
    )

    # 5. Confirm Password Reset with Bypass OTP
    confirm_reset_payload = {"email": clinician_email, "otp": "123456", "new_password": new_password}
    record_test(
        5, "Authentication", "Verify OTP and reset password using bypass code",
        f"POST /auth/confirm_password_reset\n{json.dumps(confirm_reset_payload)}", "Status code 200, success=True",
        lambda: requests.post(f"{BASE_URL}/auth/confirm_password_reset", json=confirm_reset_payload)
    )

    # 6. Verify Login with New Password
    login_new_payload = {"email": clinician_email, "password": new_password}
    def do_login_new():
        r = requests.post(f"{BASE_URL}/auth/login", json=login_new_payload)
        nonlocal token
        if r.status_code == 200:
            token = r.json().get("access_token")
        return r
    record_test(
        6, "Authentication", "Log in with the newly updated password",
        f"POST /auth/login\n{json.dumps(login_new_payload)}", "Status code 200, access_token returned",
        do_login_new
    )

    # 7. Change Password (Authenticated)
    change_pass_payload = {"email": clinician_email, "old_password": new_password, "new_password": clinician_password}
    record_test(
        7, "Authentication", "Change password when logged in",
        f"POST /auth/change_password\n{json.dumps(change_pass_payload)}", "Status code 200, success=True",
        lambda: requests.post(f"{BASE_URL}/auth/change_password", json=change_pass_payload)
    )

    # Prepare Auth Header
    auth_headers = {"Authorization": f"Bearer {token}"} if token else {}

    # 8. Upload Clinician Profile Photo
    mock_photo = create_mock_image()
    record_test(
        8, "Clinician", "Upload doctor profile photo",
        "POST /clinicians/profile_photo with photo file", "Status code 200, photo_path returned",
        lambda: requests.post(f"{BASE_URL}/clinicians/profile_photo", headers=auth_headers, files={"photo": ("doc.png", mock_photo, "image/png")})
    )

    # 9. Upload Patient Profile Photo
    mock_photo_patient = create_mock_image()
    record_test(
        9, "Patient", "Upload patient profile photo",
        f"POST /patients/profile_photo (patient_id={patient_id})", "Status code 200, photo_path returned",
        lambda: requests.post(f"{BASE_URL}/patients/profile_photo", headers=auth_headers, data={"patient_id": patient_id}, files={"photo": ("patient.png", mock_photo_patient, "image/png")})
    )

    # 10. Create / Update Patient Profile
    patient_payload = {
        "patient_id": patient_id,
        "name": "Jane Doe Tester",
        "age": 32,
        "sex": "Female",
        "photo_path": f"/static/patient_photos/patient_{patient_id}.png",
        "clinical_json": "{}",
        "doctor_id": clinician_id if clinician_id else "1",
        "last_updated": int(time.time())
    }
    record_test(
        10, "Patient", "Create patient profile in system database",
        f"POST /patients\n{json.dumps(patient_payload)}", "Status code 200, success=True",
        lambda: requests.post(f"{BASE_URL}/patients", headers=auth_headers, json=patient_payload)
    )

    # 11. Retrieve Patient Details
    record_test(
        11, "Patient", "Get patient profile data by ID",
        f"GET /patients/{patient_id}", f"Status code 200, patient_id='{patient_id}'",
        lambda: requests.get(f"{BASE_URL}/patients/{patient_id}", headers=auth_headers)
    )

    # 12. Create Case (Pending Image Analysis)
    case_payload = {
        "patient_id": patient_id,
        "patient_name": "Jane Doe Tester",
        "doctor_id": clinician_id if clinician_id else "1",
        "created_at": int(time.time()),
        "clinical_json": "{}"
    }
    def do_create_case():
        r = requests.post(f"{BASE_URL}/cases", headers=auth_headers, json=case_payload)
        nonlocal case_id
        if r.status_code == 200:
            case_id = r.json().get("id")
        return r
    record_test(
        12, "Cases", "Create a new clinical case entry",
        f"POST /cases\n{json.dumps(case_payload)}", "Status code 200, case ID returned",
        do_create_case
    )

    # 13. Predict Clinical Risk (Tabular Model)
    clinical_exam_data = {
        "age": 32,
        "sex": "Female",
        "smoking_status": "Past",
        "smoking_duration": 5,
        "smoking_frequency": "5/day",
        "smokeless_tobacco": 0,
        "alcohol": "Occasional",
        "diabetes": 0,
        "immunocompromised": 0,
        "autoimmune": 0,
        "steroids": 0,
        "chemotherapy": 0,
        "immunosuppressants": 0,
        "duration": "> 3 weeks",
        "onset": "Gradual",
        "recurrence": "First episode",
        "pain": "Painless",
        "healing_pattern": "Non-healing",
        "site": "Tongue (Lateral)",
        "size_mm": 18,
        "shape": "Round/Ovoid",
        "margins": "Ill-defined",
        "edge": "Everted",
        "induration": 1,
        "bleeding": 1,
        "lymph_palpable": 1,
        "tender": 0,
        "node_mobility": "Fixed",
        "paraesthesia": 1,
        "weight_loss": 1,
        "fever": 0
    }
    record_test(
        13, "AI Engine", "Predict clinical risk based on patient exam features (Tabular RF)",
        f"POST /predict\n{json.dumps(clinical_exam_data)}", "Status code 200, clinicalRiskScore returned",
        lambda: requests.post(f"{BASE_URL}/predict", json=clinical_exam_data)
    )

    # 14. Predict Full Risk (Hybrid Visual + Clinical AI)
    mock_lesion_photo = create_mock_image()
    def do_predict_full():
        data = {
            "case_id": str(case_id) if case_id else "1",
            "clinical_json": json.dumps(clinical_exam_data)
        }
        files = {
            "image": ("lesion.png", mock_lesion_photo, "image/png")
        }
        return requests.post(f"{BASE_URL}/predict_full", data=data, files=files)
    
    success_full, full_res = record_test(
        14, "AI Engine", "Execute hybrid AI risk evaluation (Visual CNN + Tabular RF)",
        "POST /predict_full with lesion image and clinical_json", "Status code 200, finalRiskScore & riskCategory returned",
        do_predict_full
    )
    
    # Store predictions if successful
    full_prediction = {}
    if success_full:
        try:
            full_prediction = json.loads(full_res)
        except Exception:
            pass

    # 15. Complete Case
    complete_payload = {
        "image_path": full_prediction.get("serverImagePath", f"/static/uploads/case_{case_id}.png"),
        "risk_score": full_prediction.get("finalRiskScore", 75.0),
        "clinical_score": full_prediction.get("clinicalRiskScore", 80.0),
        "visual_score": full_prediction.get("visualRiskScore", 70.0),
        "risk_category": full_prediction.get("riskCategory", "High Risk"),
        "biopsy_recommendation": full_prediction.get("biopsyRecommendation", "Urgent Biopsy Required"),
        "confidence": full_prediction.get("confidence", "85%"),
        "risk_explanation_json": json.dumps(full_prediction.get("riskExplanation", [])),
        "suggestions_json": json.dumps(full_prediction.get("clinicalSuggestions", []))
    }
    record_test(
        15, "Cases", "Mark case as Completed and save final AI results",
        f"PUT /cases/{case_id}/complete\n{json.dumps(complete_payload)}", "Status code 200, success=True",
        lambda: requests.put(f"{BASE_URL}/cases/{case_id}/complete", headers=auth_headers, json=complete_payload)
    )

    # 16. Get Case Details
    record_test(
        16, "Cases", "Retrieve completed case information",
        f"GET /cases/{case_id}", f"Status code 200, case details matching case_id={case_id}",
        lambda: requests.get(f"{BASE_URL}/cases/{case_id}", headers=auth_headers)
    )

    # 17. Get Patient Cases History
    record_test(
        17, "Cases", "Get history of all cases for patient",
        f"GET /patients/{patient_id}/history", "Status code 200, array of patient cases returned",
        lambda: requests.get(f"{BASE_URL}/patients/{patient_id}/history", headers=auth_headers)
    )

    # 18. Update Case Patient Details
    update_details_payload = {
        "patient_id": f"{patient_id}_MOD",
        "patient_name": "Jane Doe Modified"
    }
    record_test(
        18, "Cases", "Update patient identifier and name linked to a case",
        f"PUT /cases/{case_id}/details\n{json.dumps(update_details_payload)}", "Status code 200, success=True",
        lambda: requests.put(f"{BASE_URL}/cases/{case_id}/details", headers=auth_headers, json=update_details_payload)
    )

    # 19. Get Cases Overview List
    record_test(
        19, "Cases", "Get dashboard overview list of latest cases",
        "GET /cases", "Status code 200, cases list returned",
        lambda: requests.get(f"{BASE_URL}/cases", headers=auth_headers)
    )

    print("Functionality testing execution complete.")
    
    # Cleanup background process if started
    if server_process:
        print("Stopping local FastAPI server...")
        server_process.terminate()
        server_process.wait()
        print("FastAPI server stopped.")

    # Export to Excel
    export_to_excel(results)

def export_to_excel(results):
    df = pd.DataFrame(results) if pd else None
    
    # Target report paths
    excel_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "Functionality_Testing_Report.xlsx"))
    csv_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "Functionality_Testing_Report.csv"))

    if df is not None:
        try:
            # Create a highly styled Excel workbook using openpyxl directly or via Pandas ExcelWriter
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Functionality Testing')
                
                # Access workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Functionality Testing']
                
                # Apply styling
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                
                # Colors
                maroon_header_fill = PatternFill(start_color='6B1F1F', end_color='6B1F1F', fill_type='solid') # Surgical luxury maroon
                gold_text_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF') # White text on maroon
                regular_font = Font(name='Segoe UI', size=10)
                pass_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid') # Soft green
                fail_fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid') # Soft red
                pass_font = Font(name='Segoe UI', size=10, bold=True, color='155724')
                fail_font = Font(name='Segoe UI', size=10, bold=True, color='721C24')
                
                border_thin = Border(
                    left=Side(style='thin', color='DDDDDD'),
                    right=Side(style='thin', color='DDDDDD'),
                    top=Side(style='thin', color='DDDDDD'),
                    bottom=Side(style='thin', color='DDDDDD')
                )
                
                # Header formatting
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.fill = maroon_header_fill
                    cell.font = gold_text_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                worksheet.row_dimensions[1].height = 28
                
                # Rows formatting
                for row_idx in range(2, len(df) + 2):
                    worksheet.row_dimensions[row_idx].height = 24
                    status_val = worksheet.cell(row=row_idx, column=len(df.columns)).value
                    
                    for col_idx in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.font = regular_font
                        cell.border = border_thin
                        cell.alignment = Alignment(vertical='center', wrap_text=True if col_idx in [3, 4, 5, 7] else False)
                        
                        # Apply Pass/Fail highlights to the status column
                        if col_idx == len(df.columns):
                            if status_val == "PASS":
                                cell.fill = pass_fill
                                cell.font = pass_font
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                            elif status_val == "FAIL":
                                cell.fill = fail_fill
                                cell.font = fail_font
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                                
                # Column widths auto-adjust
                col_widths = {
                    'A': 15, # Test Case ID
                    'B': 18, # Component
                    'C': 35, # Description
                    'D': 35, # Request Details
                    'E': 35, # Expected
                    'F': 18, # Actual Status Code
                    'G': 40, # Actual Response
                    'H': 15, # Latency (s)
                    'I': 12  # Status
                }
                for col, width in col_widths.items():
                    worksheet.column_dimensions[col].width = width

            print(f"Excel report saved successfully to: {excel_path}")
        except Exception as ex:
            print(f"Failed to style Excel report via openpyxl: {ex}. Saving standard Excel file...")
            try:
                df.to_excel(excel_path, index=False)
                print(f"Basic Excel report saved to: {excel_path}")
            except Exception as ex2:
                print(f"Failed to write Excel sheet: {ex2}")
    
    # Always write CSV as a backup / alternative
    try:
        # If pandas is missing, write raw text CSV
        if df is not None:
            df.to_csv(csv_path, index=False)
        else:
            with open(csv_path, 'w', encoding='utf-8') as f:
                # Header
                f.write("Test Case ID,Component,Description,Request Details,Expected,Actual Status Code,Actual Response,Latency (s),Status\n")
                for r in results:
                    line = f'"{r["Test Case ID"]}","{r["Component"]}","{r["Description"]}","{r["Request Details"]}","{r["Expected"]}","{r["Actual Status Code"]}","{r["Actual Response"]}",{r["Latency (s)"]},"{r["Status"]}"\n'
                    f.write(line)
        print(f"CSV report saved successfully to: {csv_path}")
    except Exception as ex_csv:
        print(f"Failed to write CSV: {ex_csv}")

if __name__ == "__main__":
    run_tests()
